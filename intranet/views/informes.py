"""Muestra las diferentes vistas de los informes de actividades."""
import io
import os
from datetime import datetime, timedelta
from django.shortcuts import render
from django.http import JsonResponse, HttpResponse
from django.contrib.auth.models import User
from django.contrib.auth.decorators import login_required
from django.db.models import Q
# Funciones de agregación de Django
from django.conf import settings
from django.db.models import Count, Max
from django.db.models.functions import TruncMonth
from rest_framework.views import APIView
from rest_framework.response import Response
from openpyxl import Workbook
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
# Utilidades de ReportLab
from reportlab.pdfgen import canvas
from reportlab.lib.pagesizes import A4
from reportlab.lib.units import cm
from reportlab.lib import colors
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle
from auth.models import Colaborador, Area
from intranet.models import Expedientes, Ot
from kanban.models import Actividades, Tarea
from kanban.forms import ActividadesForm


def resumen(request):
    """ Informes Resumen """
    return render(request, 'informes/resumen.html')


@login_required
def detallado(request):
    """ Informes Detallados """
    form = ActividadesForm()
    context = {
        'form': form
    }
    return render(request, 'informes/detallado.html', context)


def export_actividades_excel(request):
    """Genera y devuelve un archivo Excel con todas las actividades."""
    start_date = request.GET.get('start_date', '')
    end_date = request.GET.get('end_date', '')
    search_value = request.GET.get('search_value', '').strip()

    # Filtrar actividades
    actividades = Actividades.objects.all().order_by('-id')
    if search_value:
        actividades = actividades.filter(
            Q(descripcion__icontains=search_value) |
            Q(user__username__icontains=search_value) |
            Q(ot__nombre__icontains=search_value) |
            Q(ot__id__icontains=search_value)
        )

    if start_date and end_date:
        actividades = actividades.filter(
            inicio__date__range=(start_date, end_date)
        )

    # Crear el archivo Excel
    wb = Workbook()
    ws = wb.active
    ws.title = "Actividades"

    # Encabezados
    headers = [
        'ID OT', 'Nombre OT', 'Estado', 'Actividad', 'Descripción',
        'Nombre', 'Fecha', 'Total', 'Total Decimal'
    ]

    # Insertar encabezados en la primera fila
    ws.append(headers)

    # Insertar datos
    for actividad in actividades:
        expediente = Expedientes.objects.filter(
            ot=actividad.ot).order_by('-id').first()

        ws.append([
            actividad.ot.id if actividad.ot else None,
            actividad.ot.nombre if actividad.ot else None,
            expediente.estado if expediente else None,
            actividad.tarea.titulo if actividad.tarea else actividad.descripcion,
            actividad.comentario,
            actividad.user.username if actividad.user else None,
            actividad.inicio.strftime(
                '%d/%m/%Y') if actividad.inicio else None,
            actividad.total(),
            actividad.total_decimal()
        ])

    # Definir el rango de la tabla
    tabla_rango = f"A1:I{len(actividades) + 1}"
    tabla = Table(displayName="TablaActividades", ref=tabla_rango)

    # Aplicar estilo a la tabla
    estilo = TableStyleInfo(
        name="TableStyleLight9",
        showFirstColumn=False,
        showLastColumn=False,
        showRowStripes=True,
        showColumnStripes=False
    )
    tabla = Table(ref=tabla_rango)
    tabla.displayName = "TablaActividades"
    tabla.tableStyleInfo = estilo
    for row in range(1, ws.max_row + 1):  # Recorre todas las filas existentes
        ws.row_dimensions[row].height = 25

    # Agregar la tabla a la hoja de cálculo
    ws.add_table(tabla)

    # Ajustar el ancho de las columnas automáticamente
    for col in ws.columns:
        max_length = 0
        column = col[0].column_letter  # Obtener la letra de la columna
        for cell in col:
            try:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            except Exception:
                pass
        ws.column_dimensions[column].width = max_length + 2

    # Forzar que la columna de "Descripción" tenga un ancho de 50
    col_descripcion_index = headers.index(
        "Descripción") + 1  # Obtener el índice de la columna
    col_descripcion_letter = ws.cell(
        row=1, column=col_descripcion_index).column_letter
    ws.column_dimensions[col_descripcion_letter].width = 50
    # Determinar la última fila después de insertar los datos
    # +2 porque la primera fila es de encabezados
    ultima_fila = len(actividades) + 2

    # Agregar la fórmula en la columna "Total Decimal"
    col_total_decimal_index = headers.index(
        "Total Decimal") + 1  # Índice de la columna
    col_total_decimal_letter = ws.cell(
        row=1, column=col_total_decimal_index).column_letter

    # Escribir la fórmula en la última fila
    ws[f"{col_total_decimal_letter}{ultima_fila}"] = f"=SUBTOTALES(9, {col_total_decimal_letter}2:{col_total_decimal_letter}{ultima_fila-1})"
    # Configurar respuesta HTTP
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'attachment; filename=Informes_Detallado_' + \
        search_value+'_'+start_date+'_'+end_date+'.xlsx'

    wb.save(response)
    return response


def export_resumen_excel(request):
    start_date = request.GET.get('start_date', '')
    end_date = request.GET.get('end_date', '')
    search_value = request.GET.get('search_value', '').strip()

    if not search_value:
        return JsonResponse({'error': 'El nombre del colaborador no puede estar vacío.'}, status=400)

    actividades = Actividades.objects.all().order_by('-id')
    user = User.objects.filter(username__icontains=search_value).first()
    colaborador = Colaborador.objects.filter(id_user=user).first()
    if colaborador:
        actividades = actividades.filter(user=user.id)
    else:
        return JsonResponse({'error': 'Colaborador no válido.'}, status=400)
    # Obtener el área del colaborador
    if colaborador:  # Verifica si tiene un área asignada
        area = Area.objects.filter(id=colaborador.id_are.id).first()
        # Si no se encuentra el área, poner "GENERAL"
        nombre = area.nombre.upper() if area else 'GENERAL'
    else:
        nombre = 'GENERAL'  # Si el área es NULL, asignar GENERAL

    if start_date and end_date:
        actividades = actividades.filter(
            inicio__date__range=(start_date, end_date))

    actividades_dict = {}
    for actividad in actividades:
        if actividad.ot not in actividades_dict:
            actividades_dict[actividad.ot] = []
        actividades_dict[actividad.ot].append(actividad)

    wb = Workbook()
    ws = wb.active
    ws.title = "Reporte Semanal"

    ws.merge_cells('A1:H1')
    title_cell = ws['A1']
    title_cell.value = f"REPORTE LABORAL SEMANAL DEL {start_date} AL {end_date} -DEL \nPERSONAL {search_value.upper()}  - ÁREA {nombre}"
    title_cell.font = Font(size=14, bold=True)
    title_cell.alignment = Alignment(
        horizontal='center', vertical='center', wrap_text=True)
    title_cell.fill = PatternFill(
        start_color="B0DD7F", end_color="B0DD7F", fill_type="solid")
    ws.row_dimensions[1].height = 50  # Ajusta la altura según lo necesites

    # Encabezados
    headers = ['OT', 'Nombre OT', 'Actividad', 'Responsable', 'Fechas de Elaboración',
               'Días Laborados conforme a lo delegado en sus funciones',
               'Cantidad de OT Trabajado (SEMANAL)', 'Horas SEMANAL Laborados Por OT']

    header_fill = PatternFill(start_color="C5D9F1",
                              end_color="C5D9F1", fill_type="solid")

    ws.append(headers)  # Agregar los encabezados a la hoja
    # Definir estilo de bordes negros
    thin_border = Border(
        left=Side(style='thin', color='000000'),
        right=Side(style='thin', color='000000'),
        top=Side(style='thin', color='000000'),
        bottom=Side(style='thin', color='000000')
    )
    # Aplicar alineación, negrita y color de fondo a los encabezados
    for cell in ws[2]:  # La fila 2 contiene los encabezados
        cell.alignment = Alignment(
            horizontal='center', vertical='center', wrap_text=True)
        cell.font = Font(bold=True)
        cell.fill = header_fill  # Aplicar color de fondo

    column_widths = {
        'A': 9,  # OT
        'B': 25,  # Nombre OT
        'C': 20,  # Actividad
        'D': 13,  # Responsable
        'E': 15,  # Fechas
        'F': 10,  # Días Laborados
        'G': 11,  # OT Trabajado (Semanal)
        'H': 15,  # Horas SEMANAL
    }

    # Aplicar los anchos personalizados a cada columna
    for col, width in column_widths.items():
        ws.column_dimensions[col].width = width

    for ot, actividades_ot in actividades_dict.items():
        actividades_texto = "\n".join(
            set(act.tarea.titulo for act in actividades_ot if act.tarea))
        fechas_texto = "\n".join(
            sorted(set(act.inicio.strftime('%d/%m/%Y') for act in actividades_ot)))
        dias_laborados = len(set(act.inicio.date() for act in actividades_ot))
        horas_totales = sum(
            (act.fin - act.inicio for act in actividades_ot if act.inicio and act.fin), timedelta())
        # Convertir timedelta a segundos
        total_seconds = int(horas_totales.total_seconds())
        horas_totales_decimal = horas_totales.total_seconds(
        ) / (24 * 3600)  # Convertir segundos a fracción del día
        row = [
            ot.id if ot else None,
            ot.nombre if ot else None,
            actividades_texto,
            actividades_ot[0].user.username if actividades_ot[0].user else None,
            fechas_texto,
            dias_laborados,
            1,  # OT Trabajado (Semanal) debe contar solo una vez por OT
            horas_totales_decimal  # Insertar valor numérico en lugar de cadena de texto
        ]

        ws.append(row)

        # Aplicar alineación a todas las celdas de la fila recién agregada
        for cell in ws[ws.max_row]:
            cell.alignment = Alignment(
                horizontal='center', vertical='center', wrap_text=True)
    # Configurar el idioma a español
    # Para sistemas Linux/macOS
    # locale.setlocale(locale.LC_TIME, 'Spanish_Spain.1252')  # Para Windows

    # Convertir las fechas a formato día y mes en español
    start_date_obj = datetime.strptime(start_date, '%Y-%m-%d')
    end_date_obj = datetime.strptime(end_date, '%Y-%m-%d')

    meses_es = {
        "January": "ENERO", "February": "FEBRERO", "March": "MARZO",
        "April": "ABRIL", "May": "MAYO", "June": "JUNIO",
        "July": "JULIO", "August": "AGOSTO", "September": "SEPTIEMBRE",
        "October": "OCTUBRE", "November": "NOVIEMBRE", "December": "DICIEMBRE"
    }

    # Extraer día, mes y año en el formato correcto
    dia_inicio = start_date_obj.strftime('%d')
    dia_fin = end_date_obj.strftime('%d')
    mes_nombre = meses_es[start_date_obj.strftime('%B')]
    mes_nombre1 = meses_es[end_date_obj.strftime('%B')]
    anio = start_date_obj.strftime('%Y')

   # Determinar la fila donde irá la fila total
    total_row = ws.max_row + 1

    # Combinar las celdas de A a F para el texto descriptivo
    ws.merge_cells(start_row=total_row, start_column=1,
                   end_row=total_row, end_column=6)
    total_cell = ws.cell(row=total_row, column=1)
    total_cell.value = f"TOTAL DE OTS TRABAJADOS DEL {dia_inicio} {mes_nombre}  AL {dia_fin} {mes_nombre1} {anio}"
    total_cell.font = Font(size=12, bold=True)
    total_cell.alignment = Alignment(
        horizontal='center', vertical='center', wrap_text=True)

    # Calcular la celda de la columna donde está el conteo de OTs trabajados
    # La columna donde están los valores de "OT Trabajado (Semanal)"
    columna_ot_trabajado = 7

    # Crear la fórmula para sumar todos los valores de OT trabajados (sin contar los títulos)
    inicio_datos = 3  # Primera fila con datos (después del encabezado)
    fin_datos = total_row - 1  # Última fila con datos antes del total
    formula_suma = f"=SUM(G{inicio_datos}:G{fin_datos})"

    # Insertar la suma en la celda a la derecha del total
    total_ot_cell = ws.cell(row=total_row, column=columna_ot_trabajado)
    total_ot_cell.value = formula_suma
    total_ot_cell.font = Font(size=12, bold=True)
    total_ot_cell.alignment = Alignment(
        horizontal='center', vertical='center', wrap_text=True)

    # Calcular la fila donde se insertará el nuevo texto
    total_horas_row = total_row + 1  # Fila debajo del total de OTs trabajados

    # Texto a insertar con formato personalizado
    texto_total_horas = f"TOTAL HORAS ELABORACIÓN DE OTS  DEL {dia_inicio} {mes_nombre} AL {dia_fin} {mes_nombre1} {anio}"

    # Combinar celdas de A hasta G
    ws.merge_cells(start_row=total_horas_row, start_column=1,
                   end_row=total_horas_row, end_column=7)
    total_horas_cell = ws.cell(row=total_horas_row, column=1)
    total_horas_cell.value = texto_total_horas
    total_horas_cell.font = Font(size=12, bold=True)
    total_horas_cell.alignment = Alignment(
        horizontal='center', vertical='center', wrap_text=True)

    # Calcular la suma de las horas de elaboración de OTs
    # Columna donde están las horas (ajusta según corresponda)
    columna_horas = 8
    # Sumar todas las horas
    formula_suma_horas = f"=SUM(H{inicio_datos}:H{fin_datos})"

    # Insertar la suma en la celda a la derecha (columna H)
    total_horas_suma_cell = ws.cell(row=total_horas_row, column=columna_horas)
    total_horas_suma_cell.value = formula_suma_horas
    total_horas_suma_cell.font = Font(size=12, bold=True)
    total_horas_suma_cell.alignment = Alignment(
        horizontal='center', vertical='center', wrap_text=True)
    # Aplicar formato [hh]:mm:ss a toda la columna H
    for row in ws.iter_rows(min_col=8, max_col=8, min_row=3, max_row=ws.max_row):
        for cell in row:
            cell.number_format = "[hh]:mm:ss"
   # Aplicar bordes a todas las celdas de la hoja, desde los encabezados hasta la última fila
    for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=8):
        for cell in row:
            cell.border = thin_border  # Aplicar bordes negros
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    filename = f"Reporte_Laboral_{start_date}_al_{end_date}_{search_value.upper()}_{nombre}.xlsx"
    response['Content-Disposition'] = f'attachment; filename={filename}'

    wb.save(response)
    return response


def draw_header(p, width, height, title_text):
    """
    Dibuja el logo y el título en la cabecera de una página del canvas.
    """
    # Definir la ruta absoluta al logo
    # Asume que 'static' está en el BASE_DIR del proyecto
    LOGO_PATH = os.path.join(settings.BASE_DIR, 'static', 'images', 'logo.png')

    # Título
    p.setFont("Helvetica-Bold", 16)
    p.drawCentredString(width / 2.0, height - 2*cm, title_text)

    # Dibujar el Logo (manejando si no lo encuentra)
    try:
        # Dibuja el logo en la esquina superior izquierda
        # (2cm del borde izq, 3cm del borde sup)
        # Damos un alto de 1.5cm y el ancho se ajusta automáticamente
        p.drawImage(LOGO_PATH, 2*cm, height - 3*cm, height=1.5 *
                    cm, preserveAspectRatio=True, mask='auto')
    except IOError:
        # Si no se encuentra el logo, dibuja un texto de reemplazo
        p.setFont("Helvetica", 8)
        p.setFillColorRGB(0.8, 0.2, 0.2)  # Color rojo
        p.drawString(2*cm, height - 2*cm, "[Logo no econtrado]")
        p.setFillColorRGB(0, 0, 0)  # Resetear color

# --- Funciones de Generación de PDF (Actualizadas) ---


def generar_reporte_r1(response, data_r1):
    """
    Genera el PDF para el Reporte R1: Proyectos por Mes.
    """
    p = canvas.Canvas(response, pagesize=A4)
    width, height = A4  # Obtener dimensiones

    # --- MODIFICADO: Llamar a la cabecera ---
    draw_header(p, width, height, "R1: Proyectos Registrados por Mes")

    # Cabeceras de la tabla
    p.setFont("Helvetica-Bold", 12)
    p.drawString(3*cm, height - 4*cm, "Mes")
    p.drawString(10*cm, height - 4*cm, "Cantidad de Proyectos")
    p.line(3*cm, height - 4.2*cm, width - 3*cm, height - 4.2*cm)

    # Datos
    p.setFont("Helvetica", 11)
    y_position = height - 5*cm

    for item in data_r1:
        if y_position < 4*cm:  # Salto de página
            p.showPage()
            # --- MODIFICADO: Dibujar cabecera en la nueva página ---
            draw_header(p, width, height,
                        "R1: Proyectos Registrados por Mes (Cont.)")
            p.setFont("Helvetica", 11)
            y_position = height - 4*cm  # Reiniciar Y en la nueva página

        mes_str = item['month'].strftime('%Y-%m (%B)')
        count_str = str(item['count'])

        p.drawString(3*cm, y_position, mes_str)
        p.drawString(10*cm, y_position, count_str)
        y_position -= 1*cm

    p.showPage()
    p.save()
    return response


def generar_reporte_r2(response, data_r2):
    """
    Genera el PDF para el Reporte R2: Seguimiento de Proyectos.
    """
    p = canvas.Canvas(response, pagesize=A4)
    width, height = A4

    # --- MODIFICADO: Llamar a la cabecera ---
    draw_header(p, width, height, "R2: Seguimiento y Estado de Proyectos")

    y_position = height - 4*cm

    for proyecto in data_r2:
        tareas = proyecto.tarea_set.all().order_by('orden')
        fecha_fin_estimada = tareas.aggregate(
            max_vencimiento=Max('vencimiento')
        )['max_vencimiento']

        # Estimamos el espacio necesario
        espacio_necesario = 3*cm + (len(tareas) * 0.6*cm)

        if y_position < (espacio_necesario + 4*cm):  # 4cm = margen inferior
            p.showPage()
            # --- MODIFICADO: Dibujar cabecera en la nueva página ---
            draw_header(p, width, height, "R2: Seguimiento (Continuación)")
            p.setFont("Helvetica", 11)  # Resetear fuente
            y_position = height - 4*cm  # Reiniciar Y

        # --- Dibujar Proyecto ---
        p.setFont("Helvetica-Bold", 14)
        p.setFillColorRGB(0.1, 0.1, 0.4)
        p.drawString(3*cm, y_position,
                     f"Proyecto: {proyecto.nombre} (ID: {proyecto.id})")
        y_position -= 1*cm

        p.setFont("Helvetica", 11)
        p.setFillColorRGB(0, 0, 0)
        p.drawString(3.5*cm, y_position, f"Estado General: {proyecto.estado}")
        y_position -= 0.6*cm

        fecha_str = fecha_fin_estimada.strftime(
            '%Y-%m-%d') if fecha_fin_estimada else 'No definida'
        p.drawString(3.5*cm, y_position, f"Fecha Estimada Fin: {fecha_str}")
        y_position -= 0.8*cm

        # --- Dibujar Tareas (Etapas) ---
        p.setFont("Helvetica-Bold", 11)
        p.drawString(4*cm, y_position, "Etapas / Tareas del Proyecto:")
        y_position -= 0.6*cm

        if not tareas:
            p.setFont("Helvetica-Oblique", 10)
            p.drawString(4.5*cm, y_position, "- Sin tareas registradas.")
            y_position -= 0.6*cm
        else:
            p.setFont("Helvetica", 10)
            for tarea in tareas:
                if tarea.estado == 'done':
                    p.setFillColorRGB(0, 0.5, 0)
                elif tarea.estado == 'in_progress':
                    p.setFillColorRGB(0.8, 0.5, 0)
                else:
                    p.setFillColorRGB(0, 0, 0)

                p.drawString(
                    4.5*cm, y_position, f"- {tarea.titulo} (Estado: {tarea.get_estado_display()})")
                y_position -= 0.6*cm

        p.setFillColorRGB(0, 0, 0)
        p.line(3*cm, y_position, width - 3*cm, y_position)
        y_position -= 0.5*cm

    p.showPage()
    p.save()
    return response


# --- La APIView (Sin cambios) ---

class ReportePDFView(APIView):
    """
    Vista para generar reportes R1 y R2 en PDF.

    Uso:
    GET /api/reportes-pdf/?report_type=R1
    GET /api/reportes-pdf/?report_type=R2
    GET /api/reportes-pdf/?report_type=R2&ot=123  <-- NUEVO
    """
    # permission_classes = [IsAuthenticated]

    def get(self, request, *args, **kwargs):
        report_type = request.query_params.get('report_type', None)

        if report_type == 'R1':
            # ... (La lógica de R1 no cambia) ...
            data_r1 = Ot.objects.annotate(
                month=TruncMonth('inicio')
            ).values(
                'month'
            ).annotate(
                count=Count('id')
            ).order_by('month')

            response = HttpResponse(content_type='application/pdf')
            response['Content-Disposition'] = 'attachment; filename="R1_proyectos_por_mes.pdf"'

            return generar_reporte_r1(response, data_r1)

        elif report_type == 'R2':

            # --- LÓGICA MODIFICADA PARA R2 ---

            # 1. Leer el parámetro 'ot' de la URL
            ot_id = request.query_params.get('ot', None)

            # 2. Empezar el queryset base
            proyectos = Ot.objects.all().prefetch_related('tarea_set')

            # 3. Definir un nombre de archivo por defecto
            filename = "R2_seguimiento_proyectos.pdf"

            if ot_id:
                # 4a. Si se proveyó un 'ot', filtrar por ese ID
                proyectos = proyectos.filter(id=ot_id)
                filename = f"R2_seguimiento_ot_{ot_id}.pdf"

                # Buena práctica: verificar si esa OT existe
                if not proyectos.exists():
                    return Response(
                        {"error": f"La OT con id {ot_id} no fue encontrada."},
                        status=404
                    )
            else:
                # 4b. Si NO se proveyó 'ot', mantener el comportamiento original
                # (mostrar todos los proyectos activos)
                proyectos = proyectos.filter(
                    estado='Activo').order_by('inicio')
                filename = "R2_seguimiento_proyectos_activos.pdf"

            # 5. Preparar la respuesta HTTP con el nombre de archivo dinámico
            response = HttpResponse(content_type='application/pdf')
            response['Content-Disposition'] = f'attachment; filename="{filename}"'

            # 6. Generar el PDF (la función 'generar_reporte_r2' no cambia)
            return generar_reporte_r2(response, proyectos)
            # --- FIN DE LA LÓGICA MODIFICADA ---

        else:
            return Response(
                {"error": "Debe proporcionar un 'report_type' válido en los parámetros (ej. ?report_type=R1)"},
                status=400
            )
